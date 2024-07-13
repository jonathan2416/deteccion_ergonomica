from flask import Blueprint, request, render_template, session, send_from_directory, redirect, url_for
import cv2
import mediapipe as mp
import numpy as np
import os
import uuid
from datetime import datetime
import openpyxl
import matplotlib.pyplot as plt
import psycopg2

# Crear un Blueprint de Flask llamado image_app
image_app = Blueprint('image_app', __name__)

# Ruta para subir la imagen
@image_app.route('/upload_image', methods=['GET'])
def upload_image():
    session.pop('processed_image', None)
    return render_template('upload_image.html')

# Ruta para servir las imágenes procesadas
@image_app.route('/processed_images/<filename>')
def processed_image(filename):
    return send_from_directory('processed_images', filename)

# Función para detectar poses con MediaPipe en una imagen estática
def detectar_poses_imagen(image_np):
    mp_drawing = mp.solutions.drawing_utils
    mp_pose = mp.solutions.pose

    with mp_pose.Pose(static_image_mode=True, min_detection_confidence=0.5) as pose:
        image_rgb = cv2.cvtColor(image_np, cv2.COLOR_BGR2RGB)
        results = pose.process(image_rgb)

        if not results.pose_landmarks:
            return None, None

        annotated_image = image_rgb.copy()
        mp_drawing.draw_landmarks(annotated_image, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)
        annotated_image = cv2.cvtColor(annotated_image, cv2.COLOR_RGB2BGR)

        return annotated_image, results

# Funciones de evaluación para cada parte del cuerpo
def evaluar_cabeza(pose_landmarks):
    riesgo_cabeza = 1  # Nivel 1 por defecto (cabeza alineada con la columna vertebral)

    nose = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.NOSE]
    mid_spine = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_HIP]

    if nose.y < mid_spine.y - 0.05:
        riesgo_cabeza = 2  # Cabeza inclinada hacia adelante
    elif nose.y > mid_spine.y + 0.05:
        riesgo_cabeza = 3  # Cabeza inclinada hacia atrás
    elif nose.x < mid_spine.x - 0.05:
        riesgo_cabeza = 4  # Cabeza girada lateralmente a la izquierda
    elif nose.x > mid_spine.x + 0.05:
        riesgo_cabeza = 4  # Cabeza girada lateralmente a la derecha

    return riesgo_cabeza, obtener_categoria_riesgo(riesgo_cabeza)

def evaluar_cuello(pose_landmarks):
    riesgo_cuello = 1  # Nivel 1 por defecto (cuello alineado con la columna vertebral)

    nose = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.NOSE]
    mid_spine = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_HIP]

    if nose.y < mid_spine.y - 0.05:
        riesgo_cuello = 2  # Cuello inclinado hacia adelante
    elif nose.y > mid_spine.y + 0.05:
        riesgo_cuello = 3  # Cuello inclinado hacia atrás
    elif nose.x < mid_spine.x - 0.05:
        riesgo_cuello = 4  # Cuello girado lateralmente a la izquierda
    elif nose.x > mid_spine.x + 0.05:
        riesgo_cuello = 4  # Cuello girado lateralmente a la derecha

    return riesgo_cuello, obtener_categoria_riesgo(riesgo_cuello)

def evaluar_hombros(pose_landmarks):
    riesgo_hombros = 1  # Nivel 1 por defecto (hombros relajados y alineados)

    left_shoulder = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_SHOULDER]
    right_shoulder = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_SHOULDER]

    if left_shoulder.y < 0.4 and right_shoulder.y < 0.4:
        riesgo_hombros = 2  # Hombros elevados o caídos
    elif left_shoulder.y > 0.6 and right_shoulder.y > 0.6:
        riesgo_hombros = 3  # Hombros hundidos o caídos

    return riesgo_hombros, obtener_categoria_riesgo(riesgo_hombros)

def evaluar_codos(pose_landmarks):
    riesgo_codos = 1  # Nivel 1 por defecto (codos alineados)

    left_elbow = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_ELBOW]
    right_elbow = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_ELBOW]
    left_shoulder = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_SHOULDER]
    right_shoulder = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_SHOULDER]

    if left_elbow.y < left_shoulder.y and right_elbow.y < right_shoulder.y:
        riesgo_codos = 2  # Codos extendidos más allá de 90 grados
    elif left_elbow.y > left_shoulder.y and right_elbow.y > right_shoulder.y:
        riesgo_codos = 3  # Codos flexionados menos de 90 grados

    return riesgo_codos, obtener_categoria_riesgo(riesgo_codos)

def evaluar_cintura(pose_landmarks):
    riesgo_cintura = 1  # Nivel 1 por defecto (cintura alineada)

    left_hip = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_HIP]
    right_hip = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_HIP]

    if left_hip.y < 0.3 and right_hip.y < 0.3:
        riesgo_cintura = 2  # Cintura elevada o desalineada
    elif left_hip.y > 0.7 and right_hip.y > 0.7:
        riesgo_cintura = 3  # Cintura hundida o desalineada

    return riesgo_cintura, obtener_categoria_riesgo(riesgo_cintura)

def evaluar_rodillas(pose_landmarks):
    riesgo_rodillas = 1  # Nivel 1 por defecto (rodillas alineadas)

    left_knee = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_KNEE]
    right_knee = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_KNEE]
    left_hip = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.LEFT_HIP]
    right_hip = pose_landmarks.landmark[mp.solutions.pose.PoseLandmark.RIGHT_HIP]

    if left_knee.y > left_hip.y and right_knee.y > right_hip.y:
        riesgo_rodillas = 2  # Rodillas extendidas o rígidas
    elif left_knee.y < left_hip.y and right_knee.y < right_hip.y:
        riesgo_rodillas = 3  # Rodillas flexionadas excesivamente o rígidas

    return riesgo_rodillas, obtener_categoria_riesgo(riesgo_rodillas)

def obtener_categoria_riesgo(nivel_riesgo):
    if nivel_riesgo == 1:
        return "Normal"
    elif nivel_riesgo == 2:
        return "Aceptable"
    elif nivel_riesgo == 3:
        return "Mala"
    elif nivel_riesgo == 4:
        return "Crítica"
    else:
        return "Desconocido"

# Función para guardar los resultados en un archivo Excel
def guardar_resultados_excel(resultados_postura):
    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Ergonómicos"

    # Escribir encabezados
    ws['A1'] = 'Parte del Cuerpo'
    ws['B1'] = 'Nivel de Riesgo'
    ws['C1'] = 'Descripción'

    # Evaluar cada parte del cuerpo y guardar los resultados
    for parte, (nivel, categoria) in resultados_postura.items():
        ws.append([parte, nivel, categoria])

    # Generar el nombre del archivo
    excel_filename = f"resultados_ergonomicos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_filepath = os.path.join('processed_images', excel_filename)
    wb.save(excel_filepath)

    return excel_filename

# Función para crear un gráfico ergonómico
def crear_grafico_ergonomico(resultados_postura):
    partes = list(resultados_postura.keys())
    niveles = [nivel for nivel, _ in resultados_postura.values()]

    plt.figure(figsize=(10, 5))
    plt.bar(partes, niveles, color='skyblue')
    plt.xlabel('Partes del Cuerpo')
    plt.ylabel('Nivel de Riesgo')
    plt.title('Gráfico de Evaluación Ergonómica')
    plt.ylim(0, 4)
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Guardar el gráfico como imagen en la carpeta de imágenes procesadas
    grafico_filename = f"grafico_ergonomico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    grafico_filepath = os.path.join('processed_images', grafico_filename)
    plt.savefig(grafico_filepath)
    plt.close()

    return grafico_filename

# Función para guardar los resultados en la base de datos
def guardar_resultados_db(resultados_postura):
    conn = None
    try:
        conn = psycopg2.connect(
            host="localhost",
            database="ergonomic_evaluation",
            user="postgres",
            password="2311"
        )
        cur = conn.cursor()
        
        for parte, (nivel, categoria) in resultados_postura.items():
            descripcion = obtener_categoria_riesgo(nivel)
            fecha = datetime.now()
            cur.execute("INSERT INTO resultados_ergonomicos (parte_cuerpo, nivel_riesgo, categoria_riesgo, descripcion, fecha) VALUES (%s, %s, %s, %s, %s)",
                        (parte, nivel, categoria, descripcion, fecha))
        
        conn.commit()
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()

# Ruta para procesar una imagen enviada por el usuario
@image_app.route('/process_image', methods=['POST'])
def process_image():
    if 'file' not in request.files:
        return render_template('upload_image.html', error="Por favor, seleccione una imagen antes de procesar.")
        
    file = request.files['file']

    if file.filename == '':
        return render_template('upload_image.html', error="Por favor, seleccione una imagen antes de procesar.")

    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    if file.filename.split('.')[-1].lower() not in allowed_extensions:
        return render_template('upload_image.html', error="Formato de archivo no permitido")
    
    image_bytes = file.read()
    image_np = cv2.imdecode(np.frombuffer(image_bytes, np.uint8), cv2.IMREAD_COLOR)
    
    annotated_image, results = detectar_poses_imagen(image_np)
    
    if annotated_image is None:
        return render_template('upload_image.html', error="No se detectó una pose en la imagen. Por favor, intente con otra imagen.")
    
    unique_filename = str(uuid.uuid4()) + ".png"
    processed_image_path = os.path.join('processed_images', unique_filename)
    cv2.imwrite(processed_image_path, annotated_image)
    
    # Evaluar las diferentes partes del cuerpo
    resultados_postura = {
        "Cabeza": evaluar_cabeza(results.pose_landmarks),
        "Cuello": evaluar_cuello(results.pose_landmarks),
        "Hombros": evaluar_hombros(results.pose_landmarks),
        "Codos": evaluar_codos(results.pose_landmarks),
        "Cintura": evaluar_cintura(results.pose_landmarks),
        "Rodillas": evaluar_rodillas(results.pose_landmarks)
    }
    
    # Guardar resultados en la base de datos
    guardar_resultados_db(resultados_postura)
    
    # Guardar resultados en un archivo Excel
    excel_filename = guardar_resultados_excel(resultados_postura)
    
    # Crear gráfico ergonómico
    grafico_filename = crear_grafico_ergonomico(resultados_postura)

    session['processed_image'] = unique_filename
    session['resultados_postura'] = resultados_postura
    session['excel_filename'] = excel_filename
    session['grafico_filename'] = grafico_filename

    return render_template('resultados.html', image_filename=unique_filename, resultados=resultados_postura,
                        excel_filename=excel_filename, grafico_filename=grafico_filename)

# Ruta para servir los archivos Excel generados
@image_app.route('/download_excel/<filename>')
def download_excel(filename):
    return send_from_directory('processed_images', filename, as_attachment=True)

# Ruta paraservir los gráficos ergonómicos generados
@image_app.route('/download_grafico/<filename>')
def download_grafico(filename):
    return send_from_directory('processed_images', filename, as_attachment=True)

if __name__ == '__main__':
    image_app.run(debug=True)